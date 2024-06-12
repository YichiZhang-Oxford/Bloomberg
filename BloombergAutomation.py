##Import libraries
from __future__ import print_function
from __future__ import absolute_import
from collections import OrderedDict
import blpapi
from optparse import OptionParser, Option, OptionValueError
import openpyxl
import copy
import datetime as dt
import calendar
import json
import QuantLib as ql
from tqdm import tqdm
import sys, os
from openpyxl import Workbook
import openpyxl.utils as xutil
import glob
import time
import win32com.client as win32
import Tkinter as tk
import warnings
warnings.filterwarnings("ignore")

##Configeration
file_dir = os.path.dirname(__file__)
sys.path.append(file_dir)

#FileConfiguration
file_config_path = os.path.join(file_dir, "FileConfiguration.json")
with open(file_config_path) as file_config:
    file = json.load(file_config)

#CurrencySetupConfiguration
currency_setup_config_path = os.path.join(file_dir, "CurrencySetupConfiguration.json")
with open(currency_setup_config_path) as currency_setup_config:
    currency_setup = json.load(currency_setup_config)

#Set up config
config = {
    "rec": {
        "maturity_date": None,
        "coupon": None,
        "df_vec": None,
        "dc_vec": None,
        "payment_date": None
    },
    "pay": {
        "maturity_date": None,
        "coupon": None,
        "df_vec": None,
        "dc_vec": None,
        "payment_date": None
    }
}

#Save original(default) config
default_config = copy.deepcopy(config)

##List of correncies in the given day count
currency_setup_ACT365 = list(
    filter(lambda x: "ACT/365" in list(currency_setup[x].values()), [v for v in list(currency_setup.keys())]))
currency_setup_ACT365FIXED = list(
    filter(lambda x: "ACT/365.FIXED" in list(currency_setup[x].values()), [v for v in list(currency_setup.keys())]))
currency_setup_30I360 = list(
    filter(lambda x: "30I/360" in list(currency_setup[x].values()), [v for v in list(currency_setup.keys())]))
currency_setup_30U360 = list(
    filter(lambda x: "30U/360" in list(currency_setup[x].values()), [v for v in list(currency_setup.keys())]))
currency_setup_ACT360 = list(
    filter(lambda x: "ACT/360" in list(currency_setup[x].values()), [v for v in list(currency_setup.keys())]))
currency_setup_30360 = list(
    filter(lambda x: "30/360" in list(currency_setup[x].values()), [v for v in list(currency_setup.keys())]))
currency_setup_ACTACT = list(
    filter(lambda x: "ACT/ACT" in list(currency_setup[x].values()), [v for v in list(currency_setup.keys())]))
currency_setup_DU252 = list(
    filter(lambda x: "DU/252" in list(currency_setup[x].values()), [v for v in list(currency_setup.keys())]))

##Model Side

#Payment Date Section
def ql_to_datetime(d):
    """The ql_to_datetime function is to convert QuantLib datetime format into datetime

    :param d: QuantLib datetime
    :type d: QuantLib datetime
    :return: datetime
    :rtype: datetime
    """

    return dt.datetime(d.year(), d.month(), d.dayOfMonth())

def datetime_to_ql(dt):
    """The datetime_to_ql function is to convert datetime into QuantLib datetime format

    :param dt: datetime
    :type dt: datetime
    :return: QuantLib datetime
    :rtype: QuantLib datetime
    """

    return ql.Date(dt.day, dt.month, dt.year)


def get_pay_freq_info():
    """The get_pay_freq_info function is to get all pay frequency from QuantLib

    :return: ordered dictionary of all pay frequency from QuantLib
    :rtype: ordered dictionary
    """

    pay_freq_dict = OrderedDict()
    pay_freq_dict["Annual"] = ql.Period(ql.Annual)
    pay_freq_dict["SemiAnnual"] = ql.Period(ql.Semiannual)
    pay_freq_dict["Quarterly"] = ql.Period(ql.Quarterly)
    pay_freq_dict["Monthly"] = ql.Period(ql.Monthly)
    pay_freq_dict["28 Days"] = ql.Period(ql.EveryFourthWeek)
    pay_freq_dict["Weekly"] = ql.Period(ql.Weekly)

    return pay_freq_dict


def get_currency_setup_info():
    """The get_currency_setup_info is to get every country's calendar from QuantLib. Some small countries don't have
    calendar from QuanLib, United States calendar will be used by dufault

    :return: ordered dictionary of every country's calendar from QuantLib
    :rtype: ordered dictionary
    """

    currency_setup_dict = OrderedDict()
    currency_setup_dict["CAD"] = ql.Canada()
    currency_setup_dict["CNY"] = ql.China()
    currency_setup_dict["KWD"] = ql.UnitedStates()
    currency_setup_dict["RON"] = ql.Romania()
    currency_setup_dict["CHF"] = ql.Switzerland()
    currency_setup_dict["COP"] = ql.UnitedStates()
    currency_setup_dict["KZT"] = ql.UnitedStates()
    currency_setup_dict["RUB"] = ql.Russia()
    currency_setup_dict["EUR"] = ql.Germany()
    currency_setup_dict["COU"] = ql.UnitedStates()
    currency_setup_dict["MMK"] = ql.UnitedStates()
    currency_setup_dict["SAR"] = ql.SaudiArabia()
    currency_setup_dict["GBP"] = ql.UnitedKingdom()
    currency_setup_dict["CRC"] = ql.UnitedStates()
    currency_setup_dict["MNT"] = ql.UnitedStates()
    currency_setup_dict["SGD"] = ql.Singapore()
    currency_setup_dict["JPY"] = ql.Japan()
    currency_setup_dict["CZK"] = ql.CzechRepublic()
    currency_setup_dict["MUR"] = ql.UnitedStates()
    currency_setup_dict["THB"] = ql.UnitedStates()
    currency_setup_dict["SEK"] = ql.Sweden()
    currency_setup_dict["DKK"] = ql.Denmark()
    currency_setup_dict["MWK"] = ql.UnitedStates()
    currency_setup_dict["TJS"] = ql.UnitedStates()
    currency_setup_dict["USD"] = ql.UnitedStates()
    currency_setup_dict["DOP"] = ql.UnitedStates()
    currency_setup_dict["MXN"] = ql.Mexico()
    currency_setup_dict["TRY"] = ql.Turkey()
    currency_setup_dict["AED"] = ql.UnitedStates()
    currency_setup_dict["EGP"] = ql.UnitedStates()
    currency_setup_dict["MYR"] = ql.UnitedStates()
    currency_setup_dict["TWD"] = ql.Taiwan()
    currency_setup_dict["ARS"] = ql.Argentina()
    currency_setup_dict["GEL"] = ql.UnitedStates()
    currency_setup_dict["MZN"] = ql.UnitedStates()
    currency_setup_dict["TZS"] = ql.UnitedStates()
    currency_setup_dict["AUD"] = ql.Australia()
    currency_setup_dict["GHS"] = ql.UnitedStates()
    currency_setup_dict["NGN"] = ql.UnitedStates()
    currency_setup_dict["UAH"] = ql.Ukraine()
    currency_setup_dict["AZN"] = ql.UnitedStates()
    currency_setup_dict["HKD"] = ql.HongKong()
    currency_setup_dict["NOK"] = ql.Norway()
    currency_setup_dict["UGX"] = ql.UnitedStates()
    currency_setup_dict["BDT"] = ql.UnitedStates()
    currency_setup_dict["HUF"] = ql.Hungary()
    currency_setup_dict["NZD"] = ql.NewZealand()
    currency_setup_dict["UZS"] = ql.UnitedStates()
    currency_setup_dict["BGN"] = ql.UnitedStates()
    currency_setup_dict["IDR"] = ql.Indonesia()
    currency_setup_dict["OMR"] = ql.UnitedStates()
    currency_setup_dict["VND"] = ql.UnitedStates()
    currency_setup_dict["BHD"] = ql.UnitedStates()
    currency_setup_dict["ILS"] = ql.Israel()
    currency_setup_dict["PEN"] = ql.UnitedStates()
    currency_setup_dict["XAF"] = ql.UnitedStates()
    currency_setup_dict["BRL"] = ql.Brazil()
    currency_setup_dict["INR"] = ql.India()
    currency_setup_dict["PHP"] = ql.UnitedStates()
    currency_setup_dict["XOF"] = ql.UnitedStates()
    currency_setup_dict["CLF"] = ql.UnitedStates()
    currency_setup_dict["ISK"] = ql.Iceland()
    currency_setup_dict["PKR"] = ql.UnitedStates()
    currency_setup_dict["ZAR"] = ql.SouthAfrica()
    currency_setup_dict["CLP"] = ql.UnitedStates()
    currency_setup_dict["KES"] = ql.UnitedStates()
    currency_setup_dict["PLN"] = ql.Poland()
    currency_setup_dict["ZMW"] = ql.UnitedStates()
    currency_setup_dict["CNH"] = ql.UnitedStates()
    currency_setup_dict["KRW"] = ql.SouthKorea()
    currency_setup_dict["QAR"] = ql.UnitedStates()

    return currency_setup_dict


def payment_date_calc(eff_dt, mat_dt, pay_freq, currency):
    """The payment_date_calc function is to calculate payment date

    :param eff_dt: effective date
    :type eff_dt: datetime
    :param mat_dt: maturity date
    :type mat_dt: datetime
    :param pay_freq: payment frequency
    :type pay_freq: str
    :param pay_freq: currency
    :type pay_freq: str
    :return: a list of payment date
    :rtype: list
    """

    effective_date = datetime_to_ql(eff_dt)
    termination_date = datetime_to_ql(mat_dt)
    tenor = pay_freq_info[pay_freq]
    calendar = currency_setup_info[currency]

    business_convention = ql.ModifiedFollowing
    termination_business_convention = False
    date_generation = ql.DateGeneration.Backward
    end_of_month = True

    schedule = ql.Schedule(effective_date, termination_date, tenor, \
                           calendar, business_convention, termination_business_convention, \
                           date_generation, end_of_month)
    result_date_list = list(map(ql_to_datetime, list(schedule)))
    return result_date_list


def pay_date_gen(eff_dt, mat_dt, pay_freq, currency):
    """The pay_date_gen function is to generate payment date

    :param eff_dt: effective date
    :type eff_dt: datetime
    :param mat_dt: maturity date
    :type mat_dt: datetime
    :param pay_freq: payment frequency
    :type pay_freq: str
    :param pay_freq: currency
    :type pay_freq: str
    :return: a list of payment date
    :rtype: list
    """

    return payment_date_calc(eff_dt, mat_dt, pay_freq, currency)


##Interpolation of discount factors

def check_curve_load():
    """The check_curve_load function is to check whether curve is loaded or not

    :return: True (loaded) or False (unloaded)
    :rtype: bool
    """

    for i in currency_info:
        if currency_info[i]["Basis Curve Number"][1] == None:
            return False
    if config["rec"]["payment_date"] == [] or config["pay"]["payment_date"] == []:
        return False
    return True


#ACT365FIXED
def year_fraction_ACT365FIXED(start_date, end_date):
    """The year_fraction_ACT365FIXED function is to return fraction in years between start_date and end_date,
    using ACT/365.FIXED convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: fraction in years between start_date and end_date, using ACT/365.FIXED convention
    :rtype: float
    """

    return (end_date - start_date).days / 365.0

def df_interp_ACT365FIXED(dt, df_dict):
    """The df_interp_ACT365FIXED function is to interpolate discount factor in the given date, using ACT/365.FIXED convention

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given date, using ACT/365.FIXED convention
    :rtype: float
    """

    date_list = list(df_dict.keys())
    if dt <= date_list[0]:
        return df_dict[date_list[0]]
    elif dt > date_list[-1]:
        return df_dict[date_list[-1]]

    i = 0
    while i < len(date_list) - 1 and date_list[i] <= dt:
        i = i + 1
    t1 = year_fraction_ACT365FIXED(date_list[0], date_list[i - 1])
    if t1 == 0:
        r1 = df_dict[date_list[0]]
    else:
        r1 = df_dict[date_list[i - 1]]

    t2 = year_fraction_ACT365FIXED(date_list[0], date_list[i])
    r2 = df_dict[date_list[i]]
    t = year_fraction_ACT365FIXED(date_list[0], dt)
    r = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
    return r

#ACT360
def year_fraction_ACT360(start_date, end_date):
    """The year_fraction_ACT360 function is to return fraction in years between start_date and end_date,
    using ACT/360 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: fraction in years between start_date and end_date, using ACT/360 convention
    :rtype: float
    """

    return (end_date - start_date).days / 360.0

def df_interp_ACT360(dt, df_dict):
    """The df_interp_ACT360 function is to interpolate discount factor in the given date, using ACT/360 convention

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given date, using ACT/360 convention
    :rtype: float
    """

    date_list = list(df_dict.keys())
    if dt <= date_list[0]:
        return df_dict[date_list[0]]
    elif dt > date_list[-1]:
        return df_dict[date_list[-1]]

    i = 0
    while i < len(date_list) - 1 and date_list[i] <= dt:
        i = i + 1
    t1 = year_fraction_ACT360(date_list[0], date_list[i - 1])
    if t1 == 0:
        r1 = df_dict[date_list[0]]
    else:
        r1 = df_dict[date_list[i - 1]]

    t2 = year_fraction_ACT360(date_list[0], date_list[i])
    r2 = df_dict[date_list[i]]
    t = year_fraction_ACT360(date_list[0], dt)
    r = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
    return r

#30I360
def day_count_30I360(start_date, end_date):
    """The day_count_30I360 function is to return number of days between start_date and end_date, using 30I/360 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: number of days between start_date and end_date, using 30I/360 convention
    :rtype: int
    """

    if start_date.day == 31:
        d1 = 30
    else:
        d1 = start_date.day
    if end_date.day == 31 and (start_date.day == 30 or start_date.day == 31):
        d2 = 30
    else:
        d2 = end_date.day

    return 360 * (end_date.year - start_date.year) \
           + 30 * (end_date.month - start_date.month) \
           + (d2 - d1)

def year_fraction_30I360(start_date, end_date):
    """The year_fraction_30I360 function is to return fraction in years between start_date and end_date,
    using 30I/360 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: fraction in years between start_date and end_date, using 30I/360 convention
    :rtype: float
    """

    return day_count_30I360(start_date, end_date) / 360.0


def df_interp_30I360(dt, df_dict):
    """The df_interp_30I360 function is to interpolate discount factor in the given date, using 30I360 convention

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given date, using 30I360 convention
    :rtype: float
    """

    date_list = list(df_dict.keys())
    if dt <= date_list[0]:
        return df_dict[date_list[0]]
    elif dt > date_list[-1]:
        return df_dict[date_list[-1]]

    i = 0
    while i < len(date_list) - 1 and date_list[i] <= dt:
        i = i + 1
    t1 = year_fraction_30I360(date_list[0], date_list[i - 1])
    if t1 == 0:
        r1 = df_dict[date_list[0]]
    else:
        r1 = df_dict[date_list[i - 1]]

    t2 = year_fraction_30I360(date_list[0], date_list[i])
    r2 = df_dict[date_list[i]]
    t = year_fraction_30I360(date_list[0], dt)
    r = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
    return r

#30U360
def day_count_30U360(start_date, end_date):
    """The day_count_30U360 function is to return number of days between start_date and end_date, using 30U/360 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: number of days between start_date and end_date, using 30U/360 convention
    :rtype: int
    """

    if start_date.day == 31:
        d1 = 30
    else:
        d1 = start_date.day
    if end_date.day == 31:
        d2 = 30
    else:
        d2 = end_date.day
    if start_date.month == 2 and (start_date.day == 28 or start_date.day == 29):
        d1 = 30
    elif end_date.month == 2 and (end_date.day == 28 or end_date.day == 29):
        d2 = 30

    return 360 * (end_date.year - start_date.year) \
           + 30 * (end_date.month - start_date.month) \
           + (d2 - d1)


def year_fraction_30U360(start_date, end_date):
    """The year_fraction_30U360 function is to return fraction in years between start_date and end_date, using 30U/360 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: fraction in years between start_date and end_date, using 30U/360 convention
    :rtype: float
    """

    return day_count_30U360(start_date, end_date) / 360.0

def df_interp_30U360(dt, df_dict):
    """The df_interp_30U360 function is to interpolate discount factor in the given date, using 30U/360 convention

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given date, using 30U/360 convention
    :rtype: float
    """

    date_list = list(df_dict.keys())
    if dt <= date_list[0]:
        return df_dict[date_list[0]]
    elif dt > date_list[-1]:
        return df_dict[date_list[-1]]

    i = 0
    while i < len(date_list) - 1 and date_list[i] <= dt:
        i = i + 1
    t1 = year_fraction_30U360(date_list[0], date_list[i - 1])
    if t1 == 0:
        r1 = df_dict[date_list[0]]
    else:
        r1 = df_dict[date_list[i - 1]]

    t2 = year_fraction_30U360(date_list[0], date_list[i])
    r2 = df_dict[date_list[i]]
    t = year_fraction_30U360(date_list[0], dt)
    r = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
    return r

#ACT365
def year_fraction_ACT365(start_date, end_date):
    """The year_fraction_ACT365 function is to return fraction in years between start_date and end_date, using ACT/365 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: fraction in years between start_date and end_date, using ACT/365 convention
    :rtype: float
    """

    return year_fraction_ACT365FIXED(start_date, end_date)

def df_interp_ACT365(dt, df_dict):
    """The df_interp_ACT365 function is to interpolate discount factor in the given date, using ACT/365 convention

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given date, using ACT/365  convention
    :rtype: float
    """

    date_list = list(df_dict.keys())

    if dt <= date_list[0]:
        return df_dict[date_list[0]]
    elif dt > date_list[-1]:
        return df_dict[date_list[-1]]

    i = 0
    while i < len(date_list) - 1 and date_list[i] <= dt:
        i = i + 1

    t1 = year_fraction_ACT365(date_list[0], date_list[i - 1])
    if t1 == 0:
        r1 = df_dict[date_list[0]]
    else:
        r1 = df_dict[date_list[i - 1]]

    t2 = year_fraction_ACT365(date_list[0], date_list[i])
    r2 = df_dict[date_list[i]]
    t = year_fraction_ACT365(date_list[0], dt)
    r = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
    return r

#30/360
def day_count_30360(start_date, end_date):
    """The day_count_30360 function is to return number of days between start_date and end_date, using 30/360 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: number of days between start_date and end_date, using 30/360 convention
    :rtype: int
    """

    if start_date.day == 31:
        d1 = 30
    elif start_date.month == 2 and (start_date.day == 28 or start_date.day == 29):
        d1 = 30
    else:
        d1 = start_date.day

    if (start_date.month == 2 and (start_date.day == 28 or start_date.day == 29)) and (
            end_date.month == 2 and (end_date.day == 28 or end_date.day == 29)):
        d2 = 30
    elif end_date.day == 31 and (start_date.day == 30 or start_date.day == 31 or (
            start_date.month == 2 and (start_date.day == 28 or start_date.day == 29))):
        d2 = 30
    else:
        d2 = end_date.day

    return 360 * (end_date.year - start_date.year) \
           + 30 * (end_date.month - start_date.month) \
           + (d2 - d1)

def year_fraction_30360(start_date, end_date):
    """The year_fraction_30360 function is to return fraction in years between start_date and end_date, using 30/360 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: fraction in years between start_date and end_date, using 30/360 convention
    :rtype: float
    """

    return day_count_30360(start_date, end_date) / 360.0

def df_interp_30360(dt, df_dict):
    """The df_interp_30360 function is to interpolate discount factor in the given date, using 30/360 convention

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given date, using 30/360  convention
    :rtype: float
    """

    date_list = list(df_dict.keys())

    if dt <= date_list[0]:
        return df_dict[date_list[0]]
    elif dt > date_list[-1]:
        return df_dict[date_list[-1]]

    i = 0
    while i < len(date_list) - 1 and date_list[i] <= dt:
        i = i + 1

    t1 = year_fraction_30360(date_list[0], date_list[i - 1])
    if t1 == 0:
        r1 = df_dict[date_list[0]]
    else:
        r1 = df_dict[date_list[i - 1]]

    t2 = year_fraction_30360(date_list[0], date_list[i])
    r2 = df_dict[date_list[i]]
    t = year_fraction_30360(date_list[0], dt)
    r = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
    return r

#ACT/ACT
def year_fraction_ACTACT(start_date, end_date):
    """The year_fraction_ACTACT function is to return fraction in years between start_date and end_date, using ACT/ACT convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: fraction in years between start_date and end_date, using ACT/ACT convention
    :rtype: float
    """

    if (start_date.year % 400 == 0) or ((start_date.year % 4 == 0) and (start_date.year % 100 != 0)):
        start_date_year = 366
    else:
        start_date_year = 365
    if (end_date.year % 400 == 0) or ((end_date.year % 4 == 0) and (end_date.year % 100 != 0)):
        end_date_year = 366
    else:
        end_date_year = 365

    if (end_date - start_date).days > 366:
        return (dt.datetime(start_date.year, 12, 31) - start_date) / start_date_year + (
                    end_date.year - start_date.year - 1) * 1 + \
               (end_date - dt.datetime(end_date.year, 1, 1)) / end_date_year
    else:
        return (dt.datetime(start_date.year, 12, 31) - start_date) / start_date_year + (
                    end_date - dt.datetime(end_date.year, 1, 1)) / end_date_year

def df_interp_ACTACT(dt, df_dict):
    """The df_interp_ACTACT function is to interpolate discount factor in the given date, using ACT/ACT convention

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given date, using ACT/ACT convention
    :rtype: float
    """

    date_list = list(df_dict.keys())

    if dt <= date_list[0]:
        return df_dict[date_list[0]]
    elif dt > date_list[-1]:
        return df_dict[date_list[-1]]

    i = 0
    while i < len(date_list) - 1 and date_list[i] <= dt:
        i = i + 1

    t1 = year_fraction_ACTACT(date_list[0], date_list[i - 1])
    if t1 == 0:
        r1 = df_dict[date_list[0]]
    else:
        r1 = df_dict[date_list[i - 1]]

    t2 = year_fraction_ACTACT(date_list[0], date_list[i])
    r2 = df_dict[date_list[i]]
    t = year_fraction_ACTACT(date_list[0], dt)
    r = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
    return r

#DU/252
def year_fraction_DU252(start_date, end_date):
    """The year_fraction_DU252 function is to return fraction in years between start_date and end_date, using DU/252 convention

    :param start_date: start date
    :type start_date: datetime
    :param end_date: end date
    :type end_date: datetime
    :return: fraction in years between start_date and end_date, using DU/252 convention
    :rtype: float
    """

    days = []
    excluded = (6, 7)
    while start_date.date() <= end_date.date():
        if start_date.isoweekday() not in excluded:
            days.append(start_date)
        start_date += dt.timedelta(days=1)
    return len(days) / 252

def df_interp_DU252(dt, df_dict):
    """The df_interp_DU252 function is to interpolate discount factor in the given date, using DU/252 convention

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given date, using DU/252 convention
    :rtype: float
    """

    date_list = list(df_dict.keys())

    if dt <= date_list[0]:
        return df_dict[date_list[0]]
    elif dt > date_list[-1]:
        return df_dict[date_list[-1]]

    i = 0
    while i < len(date_list) - 1 and date_list[i] <= dt:
        i = i + 1

    t1 = year_fraction_DU252(date_list[0], date_list[i - 1])
    if t1 == 0:
        r1 = df_dict[date_list[0]]
    else:
        r1 = df_dict[date_list[i - 1]]

    t2 = year_fraction_DU252(date_list[0], date_list[i])
    r2 = df_dict[date_list[i]]
    t = year_fraction_DU252(date_list[0], dt)
    r = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
    return r

def currency_df(currency, dt, df_dict):
    """The currency_df function is to return discount factor in the given currency

    :param currency: currency
    :type currency: str
    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :return: discount factor in the given currency
    :rtype: float
    """

    if currency in currency_setup_ACT365:
        return df_interp_ACT365(dt, df_dict)
    elif currency in currency_setup_30I360:
        return df_interp_30I360(dt, df_dict)
    elif currency in currency_setup_30U360:
        return df_interp_30U360(dt, df_dict)
    elif currency in currency_setup_ACT365FIXED:
        return df_interp_ACT365FIXED(dt, df_dict)
    elif currency in currency_setup_ACT360:
        return df_interp_ACT360(dt, df_dict)
    elif currency in currency_setup_30360:
        return df_interp_30360(dt, df_dict)
    elif currency in currency_setup_ACTACT:
        return df_interp_ACTACT(dt, df_dict)
    elif currency in currency_setup_DU252:
        return df_interp_DU252(dt, df_dict)

def df_interp(dt, df_dict, leg):
    """The df_interp function is to return discount factor in the given leg

    :param dt: date
    :type dt: datetime
    :param df_dict: ordered dictionary of discount factor
    :type df_dict: ordered dictionary
    :param leg: leg
    :type currency: str
    :return: discount factor in the given leg
    :rtype: float
    """

    if leg == "REC":
        currency = config_per_run["rec"]["currency"]
        return currency_df(currency, dt, df_dict)
    else:
        currency = config_per_run["pay"]["currency"]
        return currency_df(currency, dt, df_dict)

def df_generation():
    """The df_generation function is to generate discount factor

    :return: an ordered dictionary of discount factor
    :rtype: order dictionary
    """

    if check_curve_load():
        df_rec_curve = currency_info[config_per_run["rec"]["currency"]]["Basis Curve Number"][1]
        df_pay_curve = currency_info[config_per_run["pay"]["currency"]]["Basis Curve Number"][1]
        local_rec_dict = OrderedDict()
        local_pay_dict = OrderedDict()

        for i in config["rec"]["payment_date"]:
            local_rec_dict[i] = df_interp(i, df_rec_curve, "REC")

        config["rec"]["df_vec"] = local_rec_dict

        for j in config["pay"]["payment_date"]:
            local_pay_dict[j] = df_interp(j, df_pay_curve, "PAY")

        config["pay"]["df_vec"] = local_pay_dict

##Day Count Vector Generation
def dc_generation(day_cnt, pay_date):
    """The dc_generation function is to return a list of year fraction in the specific convention

    :param day_cnt: day count
    :type day_cnt: str
    :param pay_date: list of payment date
    :type pay_date: list
    :return: list of year fraction in the specific convention
    :rtype: list
    """

    dc_vec_result = []
    if day_cnt == "ACT/365":
        for i in range(len(pay_date))[1:]:
            dc_vec_result.append(year_fraction_ACT365(pay_date[i - 1], pay_date[i]))
    elif day_cnt == "30I/360":
        for i in range(len(pay_date))[1:]:
            dc_vec_result.append(year_fraction_30I360(pay_date[i - 1], pay_date[i]))
    elif day_cnt == "30U/360":
        for i in range(len(pay_date))[1:]:
            dc_vec_result.append(year_fraction_30U360(pay_date[i - 1], pay_date[i]))
    elif day_cnt == "ACT/365.FIXED":
        for i in range(len(pay_date))[1:]:
            dc_vec_result.append(year_fraction_ACT365FIXED(pay_date[i - 1], pay_date[i]))
    elif day_cnt == "ACT/360":
        for i in range(len(pay_date))[1:]:
            dc_vec_result.append(year_fraction_ACT360(pay_date[i - 1], pay_date[i]))
    elif day_cnt == "30/360":
        for i in range(len(pay_date))[1:]:
            dc_vec_result.append(year_fraction_30360(pay_date[i - 1], pay_date[i]))
    elif day_cnt == "ACT/ACT":
        for i in range(len(pay_date))[1:]:
            dc_vec_result.append(year_fraction_ACTACT(pay_date[i - 1], pay_date[i]))
    elif day_cnt == "DU/252":
        for i in range(len(pay_date))[1:]:
            dc_vec_result.append(year_fraction_DU252(pay_date[i - 1], pay_date[i]))

    return dc_vec_result

def payCouponCalc(rec_coupon, rec_df_vec, pay_df_vec, rec_dc_vec, pay_dc_vec):
    """The payCouponCalc function is to calculate pay coupon

    :param rec_coupon: receive coupon
    :type rec_coupon: float
    :param rec_df_vec: list of receive discount factor
    :type rec_df_vec: list
    :param pay_df_vec: list of pay discount factor
    :type pay_df_vec: list
    :param rec_dc_vec: list of receive day count
    :type rec_dc_vec: list
    :param pay_dc_vec: list of pay day count
    :type pay_dc_vec: list
    :return: pay coupon
    :rtype: float
    """

    coupon_rec_pv = (float(rec_coupon) / 100.0) * sum([x_i * y_i for x_i, y_i in zip(rec_df_vec, rec_dc_vec)])
    not_ex_rec_pv = rec_df_vec[-1] - pay_df_vec[-1]
    coupon_pay_pv = sum([x_i * y_i for x_i, y_i in zip(pay_df_vec, pay_dc_vec)])

    return (coupon_rec_pv + not_ex_rec_pv) / coupon_pay_pv


def run_pay_coupon_calc(rec_coupon, period):
    """The run_pay_coupon_calc function is to return pay coupon from above functions

    :param rec_coupon: receive coupon
    :type rec_coupon: float
    :param period: period
    :type period: int
    :return: pay coupon
    :rtype: float
    """

    global config
    config = load_maturity_date(rec_coupon, config_per_run["valuation"], period)
    pay_date_list = pay_date_gen(config_per_run["pay"]["effective_date"], \
                                 config["pay"]["maturity_date"], config_per_run["pay"]["pay_freq"],
                                 config_per_run["pay"]["currency"])
    rec_date_list = pay_date_gen(config_per_run["rec"]["effective_date"], \
                                 config["rec"]["maturity_date"], config_per_run["rec"]["pay_freq"],
                                 config_per_run["rec"]["currency"])

    config["rec"]["payment_date"] = rec_date_list[1:]
    config["pay"]["payment_date"] = pay_date_list[1:]

    config["pay"]["dc_vec"] = dc_generation(config_per_run["pay"]["day_count"], pay_date_list)
    config["rec"]["dc_vec"] = dc_generation(config_per_run["rec"]["day_count"], rec_date_list)

    df_generation()

    pay_coupon = payCouponCalc(config["rec"]["coupon"], \
                               [v for v in config["rec"]["df_vec"].values()], \
                               [u for u in config["pay"]["df_vec"].values()], \
                               config["rec"]["dc_vec"], \
                               config["pay"]["dc_vec"])
    pay_coupon = pay_coupon * 100
    config["pay"]["coupon"] = pay_coupon
    return pay_coupon

##Bloomberg Excel API Side (Not used)
# def check_sheet_data(bloomberg_data_wb, sheet_name):
#     """The check_sheet_data function is to check excel sheet data is loaded or not
#
#     :param bloomberg_data_wb: Bloomberg workbook
#     :type bloomberg_data_wb: workbook
#     :param sheet_name: sheet name
#     :type sheet_name: str
#     :return: None
#     :rtype: None
#     """
#
#     count = 0
#     while True:
#         readData = bloomberg_data_wb.Worksheets(sheet_name)
#         allData = readData.UsedRange
#         if allData.Rows.Count > 1 or allData.Columns.Count > 1:
#             break
#         time.sleep(1)
#         count += 1
#
# def load_discount_factor(bloomberg_data_address):
#     """The load_discount_factor function is to load discount factor from excel API
#
#     :param bloomberg_data_address: Bloomberg data excel file path
#     :type bloomberg_data_address: str
#     :return: None
#     :rtype: None
#     """
#
#     curve_date = wb["Main Tab"]["C2"].value.strftime('%Y%m%d')
#     workbook = xlsxwriter.Workbook(bloomberg_data_address)
#
#     for currency in list(config_curve["swap_curve"]):
#         worksheet = workbook.add_worksheet(currency)
#         query = """=BView(BCurveStrip(""" + '"' + config_curve["swap_curve"][currency][0] + '"' + ""","CurveDate",""" + \
#                 '"' + curve_date + '"' + """),"Maturity,DF.Mid","Headers=True")""".format(currency)
#         worksheet.write(0, 0, query)
#     workbook.close()
#
#
# def check_excel_data(bloomberg_data_address):
#     """The check_excel_data function is to check excel data is loaded or not
#
#     :param bloomberg_data_address: Bloomberg data excel file path
#     :type bloomberg_data_address: str
#     :return: None
#     :rtype: None
#     """
#
#     bb = 'C:/blp/API/Office Tools/BloombergUI.xla'
#     xl = win32.DispatchEx("Excel.Application")
#     xl.Workbooks.Open(bb)
#     xl.AddIns("Bloomberg Excel Tools").Installed = True
#     bloomberg_data_wb = xl.Workbooks.Open(Filename=bloomberg_data_address)  # opens workbook in readonly mode.
#
#     xl.Visible = False
#     xl.EnableEvents = False
#     xl.DisplayAlerts = False
#
#     for currency in currency_info:
#         check_sheet_data(bloomberg_data_wb, currency)
#
#     bloomberg_data_wb.Close(SaveChanges=1)
#     xl.DisplayAlerts = True
#     xl.Quit()
#     # Cleanup the com reference.
#     del xl
#
# def store_df(bloomberg_data_address):
#     """The store_df function is to store discount factor from excel
#
#     :param bloomberg_data_address: Bloomberg data excel file path
#     :type bloomberg_data_address: str
#     :return: None
#     :rtype: None
#     """
#
#     for currency in currency_info:
#         df_dict = OrderedDict()
#         currency_df = pd.read_excel(bloomberg_data_address, sheet_name=currency)
#         for i in range(len(currency_df)):
#             date = dt.datetime.strptime(str(currency_df.iat[i, 0]), "%Y-%m-%d %H:%M:%S")
#             # date = dt.datetime.strptime(str(currency_df.iat[i, 0]), "%m/%d/%Y")
#             df_dict[date] = currency_df.iat[i, 1]
#         currency_info[currency]["Basis Curve Number"][1] = df_dict


##Bloomberg Python API

#Load data
SECURITY_DATA = blpapi.Name("securityData")
SECURITY = blpapi.Name("security")
FIELD_DATA = blpapi.Name("fieldData")
FIELD_EXCEPTIONS = blpapi.Name("fieldExceptions")
FIELD_ID = blpapi.Name("fieldId")
ERROR_INFO = blpapi.Name("errorInfo")
RESPONSE_ERROR = blpapi.Name("responseError")
SESSION_TERMINATED = blpapi.Name("SessionTerminated")
CATEGORY = blpapi.Name("category")
MESSAGE = blpapi.Name("message")

def checkDateTime(option, opt, value):
    """The checkDateTime function is to check date is in the correct format

    :param value: date
    :type value: datetime
    :return: datetime
    :rtype: datetime
    """

    try:
        return dt.datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    except ValueError as ex:
        raise OptionValueError(
            "option {0}: invalid datetime value: {1} ({2})".format(
                opt, value, ex))

class ExampleOption(Option):
    TYPES = Option.TYPES + ("datetime",)
    TYPE_CHECKER = copy.copy(Option.TYPE_CHECKER)
    TYPE_CHECKER["datetime"] = checkDateTime

def parseCmdLine():
    parser = OptionParser(description="Swap Manager Scraper",
                          epilog="Notes: " +
                                 "1) All times are in GMT. " +
                                 "2) Only one security can be specified. " +
                                 "3) Only one event can be specified.",
                          option_class=ExampleOption)
    parser.add_option("-a",
                      "--ip",
                      dest="host",
                      help="server name or IP (default: %default)",
                      metavar="ipAddress",
                      default="localhost")
    parser.add_option("-p",
                      dest="port",
                      type="int",
                      help="server port (default: %default)",
                      metavar="tcpPort",
                      default=8194)
    (options, args) = parser.parse_args()

    return options

def printErrorInfo(leadingStr, errorInfo):
    print("%s%s (%s)" % (leadingStr, errorInfo.getElementAsString(CATEGORY),
                         errorInfo.getElementAsString(MESSAGE)))

def SWPMRequest(session, currency_list, curve_date):
    """The SWPMRequest function is to request discount factor in the given currency list and curve date

    :param session: session
    :type session: session
    :param currency_list: list of currency
    :type currency_list: list
    :param curve_date: curve date
    :type curve_date: datetime
    :return: None
    :rtype: None
    """

    refDataService = session.getService("//blp/refdata")
    request = refDataService.createRequest("ReferenceDataRequest")

    for currency in currency_list:
        request.append("securities", currency_info[currency]["Basis Curve Index"])
    overrides = request.getElement("overrides")
    override1 = overrides.appendElement()
    override1.setElement("fieldId", "REFERENCE_DATE")
    override1.setElement("value", curve_date)

    override2 = overrides.appendElement()
    override2.setElement("fieldId", "SW_CRV_ZR_DF_FREQ")
    override2.setElement("value", "QUARTERLY")

    request.append("fields", "SW_CRV_DISCOUNT_FACTORS")

    # print("Sending Request:", request)
    cid = session.sendRequest(request)

def processMessage(msg, currency_list):
    """The processMessage function is to load discount factor in the given currency list

    :param msg: message
    :type msg: message
    :param currency_list: list of currency
    :type currency_list: list
    :return: None
    :rtype: None
    """

    if not msg.hasElement(SECURITY_DATA):
        print("Unexpected message:")
        print(msg)
        return

    securityDataArray = msg.getElement(SECURITY_DATA)
    for securityData in securityDataArray.values():
        fieldData = securityData.getElement(FIELD_DATA)
        for field in fieldData.elements():
            if field.isValid():
                df_dict = OrderedDict()
                flag = True
                i = 0
                while flag:
                    try:
                        date_time_obj = dt.datetime. \
                            strptime((field.getValue(i)). \
                                     getElementAsString("Date"), '%Y-%m-%d')
                        df_dict[date_time_obj] \
                            = ((field.getValue(i)).getElementAsFloat("Discount Factor"))
                        i = i + 1
                    except:
                        flag = False
            else:
                print(field.name(), " is NULL")

        for currency in currency_list:
            if currency_info[currency]["Basis Curve Index"] == str(securityData.getElementAsString(SECURITY)):
                temp_currency = currency
                break

        currency_info[temp_currency]["Basis Curve Number"][1] = df_dict

        fieldExceptionArray = securityData.getElement(FIELD_EXCEPTIONS)
        for fieldException in fieldExceptionArray.values():
            errorInfo = fieldException.getElement(ERROR_INFO)
            print(errorInfo.getElementAsString("category"), ":", \
                  fieldException.getElementAsString(FIELD_ID))

def processResponseEvent(event, currency_list):
    for msg in event:
        # print(msg)
        if msg.hasElement(RESPONSE_ERROR):
            printErrorInfo("REQUEST FAILED: ", msg.getElement(RESPONSE_ERROR))
            continue
        processMessage(msg, currency_list)

def eventLoop(session, currency_list):
    done = False
    while not done:
        # nextEvent() method below is called with a timeout to let
        # the program catch Ctrl-C between arrivals of new events
        event = session.nextEvent(500000)
        if event.eventType() == blpapi.Event.PARTIAL_RESPONSE:
            print("Processing Partial Response")
            processResponseEvent(event, currency_list)
        elif event.eventType() == blpapi.Event.RESPONSE:
            # print("Processing Response")
            processResponseEvent(event, currency_list)
            done = True
        else:
            for msg in event:
                if event.eventType() == blpapi.Event.SESSION_STATUS:
                    if msg.messageType() == SESSION_TERMINATED:
                        done = True

def run_bloomberg_python_api(currency_list, curve_date):
    """The run_bloomberg_python_api function is to run Bloomberg Python API to download discount factor

    :param currency_list: list of currency
    :type currency_list: list
    :param curve_date: curve date
    :type curve_date: datetime
    :return: None
    :rtype: None
    """

    options = parseCmdLine()
    sessionOptions = blpapi.SessionOptions()
    sessionOptions.setServerHost(options.host)
    sessionOptions.setServerPort(options.port)

    # print("Connecting to %s:%s" % (options.host, options.port))
    # Create a Session
    session = blpapi.Session(sessionOptions)

    # Start a Session
    if not session.start():
        print("Failed to start session.")
        return

    try:
        # Open service to get historical data from
        if not session.openService("//blp/refdata"):
            print("Failed to open //blp/refdata")
            return

        SWPMRequest(session, currency_list, curve_date)
        # wait for events from session.
        eventLoop(session, currency_list)
    finally:
        # Stop the session
        session.stop()

##GUI

def GUI():
    """The GUI function is to pop-up a window to show the Python script is done

    :return: None
    :rtype: None
    """

    root = tk.Tk()
    root.title("Message")
    canvas = tk.Canvas(root, width=300, height=100)
    canvas.pack()
    button = tk.Button(root, text="Done!", command=root.destroy)
    canvas.create_window(150, 50, window=button)
    root.mainloop()

##Excel Side

#ALM Reporting Template

def get_main_info():
    """The get_main_info function is to get main information in the ALM Reporting Template [Main Tab]

    :return: ordered dictionary of main information
    :rtype: ordered dictionary
    """

    main_dict = OrderedDict()
    row_index = 0
    while wb["Main Tab"]["C5"].offset(row=row_index, column=0).value is not None:
        if wb["Main Tab"]["C5"].offset(row=row_index, column=0).value == "Y":
            swap_dict = OrderedDict()
            swap_name = wb["Main Tab"]["C5"].offset(row=row_index, column=-1).value
            swap_dict["Receive Currency"] = wb["Main Tab"]["C5"].offset(row=row_index, column=1).value
            swap_dict["Pay Currency"] = wb["Main Tab"]["C5"].offset(row=row_index, column=2).value
            main_dict[swap_name] = swap_dict
            row_index += 1
        else:
            row_index += 1
    return main_dict


def get_currency_info():
    """The get_currency_info function is to get currency information in the ALM Reporting Template [Discount Factor BBG]

    :return: ordered dictionary of currency information
    :rtype: ordered dictionary
    """

    currency_dict = OrderedDict()
    row_index = 0
    while wb["Discount Factor BBG"]["B4"].offset(row=row_index, column=0).value is not None:
        index_dict = OrderedDict()
        currency = wb["Discount Factor BBG"]["B4"].offset(row=row_index, column=0).value
        basis_curve_number = wb["Discount Factor BBG"]["B4"].offset(row=row_index, column=1).value
        index_dict["Basis Curve Number"] = [basis_curve_number, None]
        basis_curve_index = wb["Discount Factor BBG"]["B4"].offset(row=row_index, column=2).value + " Index"
        index_dict["Basis Curve Index"] = basis_curve_index
        currency_dict[currency] = index_dict
        row_index += 1
    return currency_dict

def get_df_info():
    """The get_df_info function is to load discount factor in the ALM Reporting Template [Discount Factor BBG]

    :return: ordered dictionary of currency information
    :rtype: ordered dictionary
    """

    tab = wb["Discount Factor BBG"]
    for i in range(len(list(currency_info))):
        df_dict = OrderedDict()
        currency = tab["H2"].offset(row=0, column=i * 3).value
        j = 0
        while isinstance(tab["H2"].offset(row=2 + j, column=i * 3 + 1).value, float):
            date = dt.datetime.strptime(str(tab["H2"].offset(row=2 + j, column=i * 3).value), "%Y-%m-%d %H:%M:%S")
            df_dict[date] = tab["H2"].offset(row=2 + j, column=i * 3 + 1).value
            j += 1
        currency_info[currency]["Basis Curve Number"][1] = df_dict
    return currency_info

def get_excel_info(main_dict):
    """The get_excel_info function is to get all swap tabs information in the ALM Reporting Template swap tabs

    :param main_dict: ordered dictionary of main information
    :type main_dict: ordered dictionary
    :return: ordered dictionary of swap information
    :rtype: ordered dictionary
    """

    master_dict = OrderedDict()
    curve_date = wb["Main Tab"]["C2"].value

    for i in list(main_dict.keys()):
        curve_per_run_local = OrderedDict()
        curve_per_run_local["curve_date"] = curve_date

        delta = ql.Period(2, ql.Days)
        val_dt = ql_to_datetime(ql.UnitedStates().advance(datetime_to_ql(curve_date), delta))
        curve_per_run_local["valuation"] = val_dt

        rec_leg = OrderedDict()
        pay_leg = OrderedDict()
        rec_leg["currency"] = main_dict[i]["Receive Currency"]
        pay_leg["currency"] = main_dict[i]["Pay Currency"]
        rec_leg["effective_date"] = val_dt
        pay_leg["effective_date"] = val_dt
        rec_leg["pay_freq"] = currency_setup[rec_leg["currency"]]["pay_freq"]
        pay_leg["pay_freq"] = currency_setup[pay_leg["currency"]]["pay_freq"]
        rec_leg["day_count"] = currency_setup[rec_leg["currency"]]["day_count"]
        pay_leg["day_count"] = currency_setup[pay_leg["currency"]]["day_count"]
        curve_per_run_local["rec"] = rec_leg
        curve_per_run_local["pay"] = pay_leg

        master_dict[i] = curve_per_run_local

    return master_dict

def load_maturity_date(rec_coupon, val_dt, term):
    """The load_maturity_date is to calculate maturity date

    :param rec_coupon: receive coupon
    :type rec_coupon: float
    :param val_dt: valuation date
    :type val_dt: datetime
    :param term: term
    :type term: int
    :return: ordered dictionary of local config
    :rtype: ordered dictionary
    """

    valuation_date = datetime_to_ql(val_dt)
    # mat_dt = dt.datetime.strptime(str(mat_dt), "%m%d%Y")
    term = ql.Period(term, ql.Years)
    mat_dt = ql_to_datetime(ql.UnitedStates().advance(valuation_date, term))
    # mat_dt = dt.datetime.strptime(str(ql_to_datetime(mat_dt)), "%m%d%Y")
    config_local = copy.deepcopy(default_config)
    config_local["rec"]["maturity_date"] = mat_dt
    config_local["rec"]["coupon"] = rec_coupon
    config_local["pay"]["maturity_date"] = mat_dt

    return config_local


def run_report_template():
    """The run_report_template function is to run ALM Reporting Template

    :return: None
    :rtype: None
    """

    print(file["file_name"].replace(".xlsx", "") + " is running...")
    print("\n")

    excel_data_address = file["file_path"] + file["file_name"]
    global wb
    wb = openpyxl.load_workbook(excel_data_address, data_only=True, keep_links=True)  # , keep_vba= True)
    main_info = get_main_info()
    global currency_info
    currency_info = get_currency_info()
    global pay_freq_info
    pay_freq_info = get_pay_freq_info()
    global currency_setup_info
    currency_setup_info = get_currency_setup_info()

    excel_info = get_excel_info(main_info)
    get_df_info()

    ### Bloomberg API
    if file["version"].lower() == "office":
        if len(file["override_currency"]) > 0:
            run_bloomberg_python_api(file["override_currency"], wb["Main Tab"]["C2"].value.strftime('%Y%m%d'))
        else:
            pass

    print("Bloomberg data is loaded!")
    print("\n")

    for tab in tqdm(excel_info):
        print("\n")
        print(tab + " is running...")
        global config_per_run
        config_per_run = excel_info[tab]

        table_index = 0
        while wb[tab]["E5"].offset(row=table_index, column=0).value is not None:
            term_list = [1, 2, 3, 5, 7, 10, 20, 30]
            for term_index in range(len(term_list)):
                rec_coupon = wb[tab]["E5"].offset(row=table_index + term_index, column=0).value
                period = term_list[term_index]

                pay_coupon = run_pay_coupon_calc(rec_coupon, period)

                wb[tab]["E5"].offset(row=table_index + term_index, column=1).value = pay_coupon
                gov_yield = wb[tab]["E5"].offset(row=table_index + term_index, column=2).value
                spread = 100 * (pay_coupon - gov_yield)
                wb[tab]["E5"].offset(row=table_index + term_index, column=3).value = spread
            table_index += 10

        print(tab + " is done!")

    print("\n")
    print("Saving...")
    wb.save(file["file_path"] + file["file_name"].replace(" Template.xlsx", "_") + (
        wb["Main Tab"]["C2"].value).strftime('%Y%m%d') + ".xlsx")
    wb.close()
    print("\n")
    print("File is saved!")
    print("\n")
    print(file["file_name"].replace(".xlsx", "") + " is done!")
    print("\n")
    print("Have a good day!")
    print("\n")
    print("Copyright 2020 Yichi Zhang")
    print("\n")
    GUI()

#ADB All-in Template

def get_date_info():
    """The get_date_info function is to get date information in the ADB All-in Template [Dates]

    :return: ordered dictionary of date
    :rtype: ordered dictionary
    """

    date_dict = OrderedDict()
    row_index = 0
    while wb["Dates"]["C14"].offset(row=row_index, column=0).value is not None:
        TMY_date = wb["Dates"]["C14"].offset(row=row_index, column=0).value.strftime('%Y%m%d')
        curve_date = wb["Dates"]["C14"].offset(row=row_index, column=2).value
        date_dict[TMY_date] = curve_date
        row_index += 1
    return date_dict

def get_allIn_excel_info():
    """The get_allIn_excel_info function is to load swap information in the ADB All-in Template

    :return: ordered dictionary of swap information
    :rtype: ordered dictionary
    """

    master_dict = OrderedDict()

    for i in list(date_info.keys()):
        curve_per_run_local = OrderedDict()
        curve_per_run_local["curve_date"] = date_info[i]

        delta = ql.Period(2, ql.Days)
        val_dt = ql_to_datetime(ql.UnitedStates().advance(datetime_to_ql(date_info[i]), delta))
        curve_per_run_local["valuation"] = val_dt

        rec_leg = OrderedDict()
        pay_leg = OrderedDict()
        rec_leg["currency"] = wb["Dates"]["H9"].value
        pay_leg["currency"] = wb["Dates"]["I9"].value
        rec_leg["effective_date"] = val_dt
        pay_leg["effective_date"] = val_dt
        rec_leg["pay_freq"] = currency_setup[rec_leg["currency"]]["pay_freq"]
        pay_leg["pay_freq"] = currency_setup[pay_leg["currency"]]["pay_freq"]
        rec_leg["day_count"] = currency_setup[rec_leg["currency"]]["day_count"]
        pay_leg["day_count"] = currency_setup[pay_leg["currency"]]["day_count"]
        curve_per_run_local["rec"] = rec_leg
        curve_per_run_local["pay"] = pay_leg

        master_dict[i] = curve_per_run_local

    return master_dict

def get_allIn_df_info(curve_date_index):
    """The get_allIn_df_info function is to load discount factor in the ADB All-in Template [Discount Factor BBG]

    :param curve_date_index: curve date index
    :type curve_date_index: int
    :return: ordered dictionary of currency information
    :rtype: ordered dictionary
    """

    tab = wb["Discount Factor BBG"]

    rec_df_dict = OrderedDict()
    rec_currency = tab["H2"].offset(row=0, column=curve_date_index * 5).value
    row_index = 0
    while isinstance(tab["H2"].offset(row=3 + row_index, column=curve_date_index * 5 + 1).value, float):
        date = dt.datetime.strptime(str(tab["H2"].offset(row=3 + row_index, column=curve_date_index * 5).value), "%Y-%m-%d %H:%M:%S")
        rec_df_dict[date] = tab["H2"].offset(row=3 + row_index, column=curve_date_index * 5 + 1).value
        row_index += 1
    currency_info[rec_currency]["Basis Curve Number"][1] = rec_df_dict

    pay_df_dict = OrderedDict()
    pay_currency = tab["J2"].offset(row=0, column=curve_date_index * 5).value
    row_index = 0
    while isinstance(tab["J2"].offset(row=3 + row_index, column=curve_date_index * 5 + 1).value, float):
        date = dt.datetime.strptime(str(tab["J2"].offset(row=3 + row_index, column=curve_date_index * 5).value), "%Y-%m-%d %H:%M:%S")
        pay_df_dict[date] = tab["J2"].offset(row=3 + row_index, column=curve_date_index * 5 + 1).value
        row_index += 1
    currency_info[pay_currency]["Basis Curve Number"][1] = pay_df_dict

    return currency_info

def run_excel_macro (file_path, file_name):
    """The run_excel_macro function is to run Excel macro by Python

    :param file_path: file path
    :type file_path: str
    :param file_name: file name
    :type file_name: str
    :return: None
    :rtype: None
    """

    xl = win32.gencache.EnsureDispatch('Excel.Application')

    xl.Visible = False
    try:
        wb = xl.Workbooks.Open(os.path.abspath(file_path+file_name))
        xl.Application.Run(file_name+"!Formula.UpdateFormulas")
        wb.Save()
        wb.Close()
    except:
        pass
    xl.DisplayAlerts = True
    xl.Application.Quit()
    del xl

def run_all_in_template():
    """The run_all_in_template function is to run ADB All-in Template

    :return: None
    :rtype: None
    """

    print(file["template_file_name"].replace(".xlsm", "") + " Template is running...")
    print("\n")

    print("Generate " + file["template_file_name"].replace(".xlsm", "") + " Template...")
    print("\n")

    try:
        os.remove(file["input_file_name"])
    except OSError:
        pass

    run_excel_macro(file["file_path"], file["template_file_name"])

    print(file["input_file_name"].replace(".xlsm", "") + " Template is done!")
    print("\n")
    # Load Excel file
    excel_data_address = file["file_path"] + file["input_file_name"]
    global wb
    wb = openpyxl.load_workbook(excel_data_address, data_only=True, keep_links=True)
    global date_info
    date_info = get_date_info()
    global currency_info
    currency_info = get_currency_info()
    global pay_freq_info
    pay_freq_info = get_pay_freq_info()
    global currency_setup_info
    currency_setup_info = get_currency_setup_info()
    excel_info = get_allIn_excel_info()

    curve_date_index = 0
    for tab in excel_info:
        print(tab + " is running...")
        # print("\n")
        get_allIn_df_info(curve_date_index)
        ### Bloomberg API
        if file["version"].lower() == "office":
            if len(file["override_currency"]) > 0:
                run_bloomberg_python_api(file["override_currency"], date_info[tab].strftime('%Y%m%d'))
            else:
                pass
        print("Bloomberg data is loaded!")
        # print("\n")

        global config_per_run
        config_per_run = excel_info[tab]
        column_index = 0
        while wb[tab]["E5"].offset(row=0, column=column_index).value is not None:
            row_index = 0
            while wb[tab]["E5"].offset(row=row_index, column=column_index).value is not None:
                rec_coupon_cell = wb[tab]["E5"].offset(row=row_index, column=column_index)
                rec_coupon = rec_coupon_cell.value

                term_list = [1, 2, 3, 5, 7, 10, 20, 30]
                period = term_list[row_index]
                pay_coupon = run_pay_coupon_calc(rec_coupon, period)
                rec_coupon_cell.offset(row=0, column=1).value = pay_coupon
                gov_yield = rec_coupon_cell.offset(row=0, column=2).value
                spread = 100 * (pay_coupon - gov_yield)
                rec_coupon_cell.offset(row=0, column=3).value = spread

                row_index += 1
            column_index += 9
        curve_date_index += 1
        print(tab + " is done!")
        print("\n")
    print("Saving...")
    end_month_date = wb["Dates"]["G9"].value
    wb.save(file["file_path"] + file["input_file_name"].replace("_Input.xlsm", "") + "_" + end_month_date + ".xlsx")
    wb.close()
    print("\n")
    print("File is saved!")
    print("\n")
    print(file["input_file_name"].replace(".xlsm", "") + "_" + end_month_date + " Template is done!")
    print("\n")
    print("Have a good day!")
    print("\n")
    print("Copyright 2020 Yichi Zhang")
    print("\n")
    GUI()

## Main part
def main():
    """The main function is to choose to run ALM Reporting Template or ADB All-in Template

    :return: None
    :rtype: None
    """

    print("Hello "+ file["user"]+",")
    print("\n")
    if file["type"] == "A":
        run_report_template()
    if file["type"] == "B":
        run_all_in_template()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Ctrl+C pressed. Stopping...")
